import pandas as pd
import numpy as np
import json
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import logging

#logging.basicConfig(filename="edt.log", level=logging.DEBUG,\
      #format='%(asctime)s -- %(name)s -- %(levelname)s -- %(message)s')

# pour afficher les DataFrame en entier:

pd.set_option("display.max_rows", None)
#pd.set_option("max_columns", None)

# pour reset options:
#pd.reset_option(“max_columns”)
#pd.reset_option(“max_rows”)

plages_horraires = [111, 112, 113, 114, 121, 122, 123, 124, 211, 212, 213, 214, 221, 222, 223, 224, 311, 312, 313, 314, 411, 412, 413, 414, 421, 422, 423, 424, 511, 512, 513, 514, 521, 522, 523, 524, 611, 612, 613, 614]

def opener(xlsx):
    
    c = pd.read_excel(xlsx)

    # on remplace tous les NaN des la colonne ph par "[0,0,0]"
        
    c["ph"]= c["ph"].fillna("[0,0,0]")

    # on remplace tous les NaN de la colonne prof par " "

    c["prof"] = c["prof"].fillna(" ")

    # on remplace tous les NaN de la colonne matiere par ""

    c["matiere"] = c["matiere"].fillna("")

    # on remplace tous les NaN de la colonne semaine par ""

    c["semaine"] = c["semaine"].fillna("")

    # on remplace tous les NaN de la colonne classe par ""

    c["classe"] = c["classe"].fillna("")

    # on remplace tous les NaN de la colonne salle par ""

    c["salle"] = c["salle"].fillna("")

    return c

def block_horraires(ut, ph=plages_horraires):

    r = []

    for i in ph:
        
        k = list(range(i, i+ut))
        
        if max(k) in plages_horraires:
            
            r.append(k)

    return r

def is_include(a, b):

    """a et b sont des listes"""

    for i in a:
        for j in b:
            if i == j:
                return True

    return False

def prof_is_dispo2(c, n, *ph):

    php = list(set(c.loc[c.prof==n, "ph"].tolist()))
    
    for i in php:
        
        i = json.loads(i)
        
        if is_include(i, ph):

            return False
    
    return True
      
def hmat(c, classe, matiere, jour):

    """
    Retourne le nombre d’heure d’une matiere donnée dans une classe donnée pour un jour donné
    jour (int): 0 pour lundi, 1 pour mardi...
    """

    j = [111, 112, 113, 114, 121, 122, 123, 124]

    phj = [i + (jour*100) for i in j]

    x = c.loc[(c.classe==classe) & (c.matiere.str.contains(matiere)), "ph"].tolist()

    counter = 0

    for i in x:

        i = json.loads(i)

        for k in i:

            if k in phj:

                counter += 1

    return counter

def classe_is_dispo2(c, n , *ph):

    phc = list(set(c.loc[c.classe==n, "ph"].tolist()))

    for i in phc:

        i = json.loads(i)

        if is_include(i, ph):

            return False

    return True

def salle_is_dispo(c, n, *ph):

    if n == "":

        return True

    phs = list(set(c.loc[c.salle==n, "ph"].tolist()))

    for i in phs:

        i = json.loads(i)

        if is_include(i, ph):

            return False

    return True

def regc_is_dispo(c, regc, *ph):

    """regc est une liste de classe en regroupement"""

    if type(regc) != type([]):

        return True

    for i in regc:

        if classe_is_dispo2(c, i, *ph) == False:

            return False

    return True

def regp_is_dispo(c, regp, *ph):

    """regp est une liste de prof en regroupement"""

    if type(regp) != type([]):

        return True

    for p in regp:

        if prof_is_dispo2(c, p, *ph) == False:

            return False

    return True

def regs_is_dispo(c, regs, *ph):

    """regp est une liste de salle en regroupement"""

    if type(regs) != type([]):

        return True

    for s in regs:

        if salle_is_dispo(c, s, *ph) == False:

            return False

    return True

def get_random_block(ut, ut2placed = True):

    """choisi des block seances de durée ut au hasard.
    si ut2placed == True (default): les block de durée 2 ne sont pas placés au milieu des demie journées"""

    bh = list(block_horraires(ut))

    bh = [str(k) for k in bh]

    random_block = np.random.choice(bh, len(bh), replace=False)

    rb =  [json.loads(k) for k in random_block]

    if ut == 2 and ut2placed == True:

        for i in rb[:]: #on itere sur une copie de rb car on ne peut pas remove dans une boucle for directement sur la liste bouclée
            
            if i[0] in [112,122,212,222,312,412,422,512,522,612]:
                
                rb.remove(i)
            
    return rb

def get_reg(c):
    
    """ retourne un df des seances en regroupement (seances maitresses)"""

    r = c[(c.regroup > 0)]
    
    r = r.regroup.tolist()
    
    r  = list(set(r))

    #r = c.loc[r] c'est déprécié !

    r = c.reindex(r)

    return r

def testone(c, maxh=2):

    """
    teste un edt. maxh est le maximum d’heure par matiere et par jour, pour eviter d’avoir 3 x 1 h de Fr dans la meme journée par exemple
    """

    c.sort_values(by=['ut'], inplace=True, ascending=False)

    r = get_reg(c)

    stop = False

    for seance in r.itertuples():

        if seance.ph != "[0,0,0]":

            continue

        random_block = get_random_block(seance.ut)

        for j in random_block:

            # on verifie qu’il n’y a pas déjà davantage d’heure de la matiere dans la journée du random_block

            jour = (j[0]//100) - 1

            if hmat(c, seance.classe, seance.matiere, jour) >= maxh:

                continue

            # recherche des classes en regrouppement

            regc = c.loc[(c.regroup == seance.id), "classe"]
            
            regc = regc.values.tolist()

            regc.append(seance.classe)

            regc = list(set(regc))

            # recherche des prof en regrouppement

            regp = c.loc[(c.regroup == seance.id), "prof"]

            regp = regp.values.tolist()

            regp.append(seance.prof)

            regp = list(set(regp))

            # recherche des salle en regrouppement

            regs = c.loc[(c.regroup == seance.id), "salle"]

            regs = regs.values.tolist()

            regs.append(seance.salle)

            regs = list(set(regs))
            
            if "" in regs:

                regs.remove("")
            
            # recherche des plages horraires libres

            if regc_is_dispo(c, regc, *j) and regp_is_dispo(c, regp, *j) and regs_is_dispo(c, regs, *j):

                c.loc[(c.id == seance.id), "ph"] = json.dumps(j)
                
                c.loc[(c.regroup == seance.id), "ph"] = json.dumps(j)
                
                break

        if c.loc[(c.id == seance.id), "ph"].values == "[0,0,0]":

            stop = True

            break

    if not stop:   

        for seance in c.itertuples():

            random_block = get_random_block(seance.ut)

            if seance.ph == "[0,0,0]" and type(seance.regroup) != type(1):

                for j in random_block:

                    # on verifie qu’il n’y a pas déjà davantage d’heure de la matiere dans la journée du random_block

                    jour = (j[0]//100) - 1

                    if hmat(c, seance.classe, seance.matiere, jour) >= maxh:

                        continue

                    # recherche des plages horraires libres

                    if classe_is_dispo2(c, seance.classe, *j) and prof_is_dispo2(c, seance.prof, *j) and salle_is_dispo(c, seance.salle, *j):

                        c.loc[(c.id == seance.id), "ph"] = json.dumps(j)
                    
                        break             

                if c.loc[(c.id == seance.id), "ph"].values == "[0,0,0]":
                    
                    break

    c.sort_values(by=["id"], inplace=True)

    return c

def xlmaker(n):

    wb = Workbook()

    wbp = pd.ExcelFile("etab/{0}-output.xlsx".format(n))

    for s in wbp.sheet_names:

        ws = wb.create_sheet(title=s)

        df = pd.read_excel(wbp, s)

        df = df.fillna("") # on remplace tous les NaN par ""

        for row in dataframe_to_rows(df, index=False, header=True):
    
            ws.append(row)

        ws.insert_rows(6) # on insere une ligne pour la pause du midi
    
        for c in [1,2,3,4,5,6]: # on parse les 3 premieres heurs du matin et de l'ap sur le df
       
            for r in [0,4]: # on traite les block de 4 heures

                if df.iat[r,c] == df.iat[r+1, c] and df.iat[r,c] == df.iat[r+2, c] and df.iat[r,c] == df.iat[r+3, c]:

                    if r == 0: # il y'a un decalalge de 2 sur les lignes et de 1 sur les col entre le df et le wb openpyxl

                        ws.merge_cells(start_row=r+1+1, start_column=c+1, end_row=r+3+1+1, end_column=c+1)

                    if r == 4: # decalage de 3 sur les lignes et de 1 sur les colonne ( a cause de la ligne du midi insérée)

                        ws.merge_cells(start_row=r+1+1+1, start_column=c+1, end_row=r+3+1+1+1, end_column=c+1)

            for r in [0,1,4,5]: # on traite les block de 3 qui ne sont pas inclus dans les block de 4, on evite les depassements (if elif)
            
                if r in [0,4] and df.iat[r,c] == df.iat[r+1, c] and df.iat[r,c] == df.iat[r+2, c] and df.iat[r,c] != df.iat[r+3, c]:

                    if r in [0,1]:

                        ws.merge_cells(start_row=r+1+1, start_column=c+1, end_row=r+2+1+1, end_column=c+1)

                    if r in [4,5]:

                        ws.merge_cells(start_row=r+1+1+1, start_column=c+1, end_row=r+2+1+1+1, end_column=c+1)

                elif r in [1,5] and df.iat[r,c] == df.iat[r+1, c] and df.iat[r,c] == df.iat[r+2, c]:

                    if r in [0,1]:

                        ws.merge_cells(start_row=r+1+1, start_column=c+1, end_row=r+2+1+1, end_column=c+1)

                    if r in [4,5]:

                        ws.merge_cells(start_row=r+1+1+1, start_column=c+1, end_row=r+2+1+1+1, end_column=c+1)
   
            for r in [0,1,2,4,5,6]: # on traite les block de 2 non inclus dans les block de 3, on evite les depassements (if elif)

                if r in [0,1,4,5] and df.iat[r,c] == df.iat[r+1, c] and df.iat[r,c] != df.iat[r+2, c]:

                    if r in [0,1]:

                        ws.merge_cells(start_row=r+1+1, start_column=c+1, end_row=r+1+1+1, end_column=c+1)

                    if r in [4,5]:

                        ws.merge_cells(start_row=r+1+1+1, start_column=c+1, end_row=r+1+1+1+1, end_column=c+1)
               
                elif df.iat[r,c] == df.iat[r+1, c]:

                    if r in [0,1,2]:

                        ws.merge_cells(start_row=r+1+1, start_column=c+1, end_row=r+1+1+1, end_column=c+1)

                    if r in [4,5,6]:

                        ws.merge_cells(start_row=r+1+1+1, start_column=c+1, end_row=r+1+1+1+1, end_column=c+1)

        ws["A1"] = ""

        ws.delete_rows(11)

        # on fait l'alignement vert et hor de chaque cellule

        for col in ws.columns:
            
            col_letter = get_column_letter(col[0].column)
            
            ws.column_dimensions[col_letter].width = 20  # ajustement column size

            for cell in col:
            
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
            
                cell.alignment = alignment_obj

    wb.save("etab/{0}_EDT.xlsx".format(n))

def testx(xlsx, n="mon_etab",  mh=2):

    t1 = time.time()

    original = opener(xlsx)

    c = original.copy()

    c = testone(c , maxh=mh)

    while "[0,0,0]" in list(c.ph):
        
        c = original.copy()

        c = testone(c, maxh=mh)
    
    t2 = time.time()

    # liste des classes:

    lc = c.classe.tolist()
    lc = list(set(lc))
    
    try:
        lc.remove("")
    except:
        pass

    #nettoyage de la liste classe pour enlever les classes nan dues aux contraintes prof ajoutées
    for i in lc:
        if type(i) != type("str"):
            lc.remove(i)
    
    # creation des emplois du temps vierge pour chaque classe

    ent = {}

    for i in lc:

        d = {"1":[""]*8, "2":[""]*8, "3":[""]*8, "4":[""]*8, "5":[""]*8, "6":[""]*8}

        df = pd.DataFrame(d)

        df.index = ["11", "12", "13", "14", "21", "22", "23", "24"]

        ent[i] = df
    
    for i in range(len(c)):
    
        s = c.iloc[i]
   
        p = json.loads(s.ph)
    
        for k in p:
        
            col = str(k)[0]
        
            lin = str(k)[1:3]
            
            try:
                dff = ent[s.classe]
            
                dff.loc[[lin],[col]] += "\n" + s.matiere + " " + s.semaine + " " + s.salle + "\n" + s.prof + "\n" 

                ent[s.classe] = dff        
            
            except:
                pass

    # liste des profs:

    lp = c.prof.tolist()
    lp = list(set(lp))
    
    try:
        lp.remove(" ")
    except:
        pass

    #nettoyage de la liste prof pour enlever les nan dues aux contraintes classes ajoutées
    for i in lp:
        if type(i) != type("str"):
            lp.remove(i)

    # creation des emplois du temps vierges pour chaque prof

    entp = {}

    for i in lp:

        d = {"1":[""]*8, "2":[""]*8, "3":[""]*8, "4":[""]*8, "5":[""]*8, "6":[""]*8}

        df = pd.DataFrame(d)

        df.index = ["11", "12", "13", "14", "21", "22", "23", "24"]

        entp[i] = df

    for i in range(len(c)):
    
        s = c.iloc[i]
   
        p = json.loads(s.ph)

        for k in p:
        
            col = str(k)[0]
        
            lin = str(k)[1:3]
            
            try:
                dff = entp[s.prof]
                
                dff.loc[[lin],[col]] += "\n" + s.classe + "\n" + s.matiere + " " + s.semaine + " " + s.salle + "\n"
                
                entp[s.prof] = dff        
            
            except:
                pass

    # renomage des colonnes et index

    for i in lp:

        entp[i].columns = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"]

        entp[i].index = ["M1", "M2", "M3", "M4", "A1", "A2", "A3", "A4"]

    for i in lc:

        ent[i].columns = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"]

        ent[i].index = ["M1", "M2", "M3", "M4", "A1", "A2", "A3", "A4"]

    # creation du html et publication sur ipfs infura

    ENT = ""

    for i in ent.keys():

        ENT += "<p>{0}</p>".format(i)

        z = ent[i].to_html()

        z = z.replace("\n", "")

        ENT += "<div>{0}</div>".format(z)

    for i in entp.keys():

        ENT += "<p>{0}</p>".format(i)

        z = entp[i].to_html()

        z = z.replace("\n", "")

        ENT += "<div>{0}</div>".format(z)

    ENT = ENT.replace("\n", "")
   
    ENT = ENT.replace("\\n", "</br>")
 
    # creation du workbook excel

    with pd.ExcelWriter('etab/{0}-output.xlsx'.format(n)) as writer:  
       
        for i in ent.keys():
            
            df = ent[i]

            df.to_excel(writer, sheet_name=i)
                                                    
        for j in entp.keys():

            df = entp[j]

            df.to_excel(writer, sheet_name=j)

    # modification du workbook avec openpyxl
    
    xlmaker(n)

    return (c, ent, entp, ENT, t2-t1)
